# -*- coding: utf-8 -*-
"""Журнал выгрузок <HISTORY_NAME>.xlsx.

Назначение:
    - Хранить историю фактов экспорта/подготовки IFC по моделям Revit.
    - Обеспечивать быстрый ответ на вопрос «актуальна ли модель».
    - Обеспечивать удобный для чтения Excel-отчёт (<HISTORY_NAME>.xlsx).

Контракты:
    - В <HISTORY_NAME>.xlsx хранится таблица с двумя колонками:
        * путь к модели (.rvt);
        * дата модификации RVT на момент выгрузки (datetime, до минут).
    - Вся логика сравнения дат работает в «минутной» точности:
        * datetime округлён до минут до записи в историю;
        * чтение <HISTORY_NAME>.xlsx возвращает datetime также «до минут».
    - Отсутствие файла <HISTORY_NAME>.xlsx трактуется как «история пустая».

Поведение:
    - При чтении:
        * строки читаются до первой полностью пустой строки;
        * строки без пути или с некорректной датой пропускаются;
        * отсутствие файла или листа истории считается пустой историей.
    - При записи:
        * лист истории полностью пересоздаётся (строки данных);
        * таблица Excel и автофильтр пересоздаются заново;
        * при отсутствии данных сохраняется скелет A1:B2.

Особенности:
    - HistoryStore поддерживает «откат» модели во времени:
        * если mtime модели меньше последней записи по этому пути,
          более поздние записи удаляются, последняя дата обновляется.
"""
import logging
from pathlib import Path
from datetime import datetime
from typing import (
    cast,
    Dict,
    List,
    Tuple,
    Optional,
    Iterable,
)

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

from config import (
    LOGGER_NAME,
    HISTORY_PATH,
    SHEET_HISTORY,
)
from config.excel import (
    HISTORY_HDR_COL1,
    HISTORY_HDR_COL2,
    HISTORY_TBL_NAME,
    HISTORY_COL_RVT_PATH,
    HISTORY_COL_DATETIME,
    FORMAT_DATETIME_EXCEL,
)

from core.models import RevitModel

from utils.fs import ensure_dir
from utils.xlsx_helpers import Xlsx

# Модульный логгер (наследует конфигурацию от "export_ifc")
log = logging.getLogger(f"{LOGGER_NAME}.history")

# Оформление листа
COL_WIDTH_PATH = 150      # ширина колонки с путями к моделям
COL_WIDTH_DATE = 40       # ширина колонки с датами выгрузки
ALIGN_HEADER = "center"   # горизонтальное выравнивание заголовков
ALIGN_CELL = "left"       # горизонтальное выравнивание обычных ячеек
HEADER_FONT_COLOR = "000000"  # чёрный цвет шрифта заголовка
HEADER_FONT_SIZE = 14         # размер шрифта заголовка (pt)

# Тип одной строки истории: (путь, дата)
HistoryRow = Tuple[str, datetime]


# ----------------------- HistoryManager (фасад) -----------------------
class HistoryManager:
    """
    Управляет историей выгрузок (точка входа для проекта).

    Состояние:
        - _store — HistoryStore (актуальные записи в памяти);
        - _io    — HistoryXlsxIO (чтение/запись Excel).

    Методы:
        - is_up_to_date(model) — проверяет наличие актуальной записи;
        - update_record(model) — обновляет состояние истории;
        - save()               — сохраняет изменения в <HISTORY_NAME>.xlsx.

    Детали:
        - При инициализации загружает текущее состояние из файла
          (если он существует).
        - Внешний код (Exporter и др.) работает только через этот фасад.
    """

    def __init__(self, history_path: Path = HISTORY_PATH) -> None:
        """Инициализация менеджера истории.

        :param history_path: Путь к файлу <HISTORY_NAME>.xlsx.
        """
        log.info("Инициализация истории выгрузок: %s", history_path)
        self._io = HistoryXlsxIO(history_path)

        # Загружаем существующие записи из файла (если он есть).
        initial_rows = self._io.load_rows()
        self._store = HistoryStore(initial_rows)

    # ----------------------------- публичный API -----------------------------
    def is_up_to_date(self, model: RevitModel) -> bool:
        """Проверяет, есть ли в истории актуальная запись по модели.

        Сравнение:
            - путь: str(model.rvt_path);
            - дата: model.last_modified (уже «до минут»).

        :param model: Экземпляр RevitModel.
        :return: True, если запись в истории совпадает с текущим состоянием.
        """
        return self._store.is_up_to_date(model)

    def update_record(self, model: RevitModel) -> None:
        """Обновляет историю по модели, учитывая возможный «откат».

        Логика:
            - если новой даты ещё нет или она больше последней — добавляем;
            - если дата совпадает с последней — ничего не делаем;
            - если дата меньше последней — удаляем более поздние записи
              по этому пути и фиксируем новую дату.

        :param model: Экземпляр RevitModel.
        """
        self._store.update_record(model)

    def save(self) -> None:
        """Сохраняет текущее состояние истории в <HISTORY_NAME>.xlsx."""
        rows = self._store.rows_sorted()
        log.info("Сохранение истории выгрузок (%d строк)", len(rows))
        self._io.save_rows(rows)
        log.info("История выгрузок сохранена в: %s", self._io.path.name)


# ----------------------- HistoryStore -----------------------
class HistoryStore:
    """Хранит и управляет записями истории выгрузок (в памяти).

    Состояние:
        - _rows — список всех записей (path, datetime);
        - _last — dict: путь → последняя (максимальная) дата;
        - _seen — set для защиты от точных дублей (path, datetime).

    Правила:
        - На один путь может быть несколько записей (история изменений).
        - Последней считается запись с максимальной датой.
        - Дубликаты (один и тот же path + datetime) игнорируются.
    """

    def __init__(
        self,
        initial_rows: Optional[Iterable[HistoryRow]] = None,
    ) -> None:
        """Инициализирует историю предыдущих выгрузок.

        :param initial_rows: Итерация пар (путь, дата) для начальной загрузки.
        """
        # Полный список строк истории (path, dt)
        self._rows: List[HistoryRow] = []
        # Последняя дата по каждому пути
        self._last: Dict[str, datetime] = {}
        # Набор для защиты от дублей (path, dt)
        self._seen: set[HistoryRow] = set()

        if initial_rows:
            for path, dt in initial_rows:
                self.add(path, dt)

    # ----------------------------- публичный API -----------------------------
    def add(self, path_str: str, dt: datetime) -> None:
        """Добавляет запись и обновляет последнюю дату по пути.

        :param path_str: Путь к модели.
        :param dt:       Дата выгрузки (нормализованная до минут).
        """
        key = (path_str, dt)

        # Защита от точных дублей (path + datetime).
        if key in self._seen:
            return

        self._rows.append(key)
        self._seen.add(key)
        self._last[path_str] = max(dt, self._last.get(path_str, dt))

    def is_up_to_date(self, model: RevitModel) -> bool:
        """True, если дата модели совпадает с последней записью в истории.

        :param model: Экземпляр RevitModel.
        :return: True, если история актуальна относительно модели.
        """
        return self._last.get(str(model.rvt_path)) == model.last_modified

    def update_record(self, model: RevitModel) -> None:
        """Обновляет историю по пути модели с учётом возможного отката.

        Логика:
            - нет записи или новая дата больше → add(path, dt);
            - новая дата равна последней → ничего не делаем;
            - новая дата меньше последней → считаем, что проект «откатили»:
                * удаляем все записи по этому пути с датой > новой;
                * добавляем новую запись (add(path, dt)).

        :param model: Экземпляр RevitModel.
        """
        path = str(model.rvt_path)
        current_dt = model.last_modified
        last_dt = self._last.get(path)

        if last_dt is None or current_dt > last_dt:
            # Первая запись или движение вперёд.
            self.add(path, current_dt)
            return

        if current_dt == last_dt:
            # Совпадение — ничего не делаем.
            return

        # Откат по времени: чистим «будущее» и фиксируем новое состояние.
        self._prune_future_records(path, current_dt)
        self.add(path, current_dt)

    # ----------------------------- внутренние -----------------------------
    def _prune_future_records(self, path: str, threshold: datetime) -> None:
        """Удаляет записи пути с датой > threshold и пересобирает индекс.

        :param path:      Путь к модели.
        :param threshold: Верхняя граница даты (текущее состояние модели).
        """
        # Оставляем только записи:
        #   - других путей;
        #   - этого пути, но с датой <= threshold.
        self._rows = [
            (p, dt) for (p, dt) in self._rows if p != path or dt <= threshold
        ]

        # Пересобираем набор seen и индекс последних дат
        self._seen = set(self._rows)
        self._reindex_last(path)

    def _reindex_last(self, path: Optional[str] = None) -> None:
        """Пересчитывает словарь последних дат (_last).

        :param path: Путь для частичного пересчёта (None → полный пересчёт).
        """
        if path is not None:
            # Частичный пересчёт по одному пути
            dates = [dt for (p, dt) in self._rows if p == path]
            if dates:
                self._last[path] = max(dates)
            else:
                self._last.pop(path, None)
            return

        # Полный пересчёт (используется редко, но оставлен для надёжности)
        self._last.clear()
        for p, dt in self._rows:
            self._last[p] = max(dt, self._last.get(p, dt))

    def rows_sorted(self) -> List[HistoryRow]:
        """Возвращает детерминированный список строк (путь ASC, дата DESC).

        Удобно для записи в Excel: сначала группировка по пути, внутри — от
        более новых записей к старым.

        :return: Отсортированный список строк истории.
        """
        return sorted(self._rows, key=lambda t: (t[0], -t[1].timestamp()))


# ----------------------- HistoryXlsxIO -----------------------
class HistoryXlsxIO:
    """Чтение и запись Excel-файла <HISTORY_NAME>.xlsx.

    Задачи:
        - загрузка строк (path, datetime) из файла;
        - полная перезапись листа истории с шапкой и таблицей;
        - оформление таблицы и автофильтра.
    """

    def __init__(self, path: Path = HISTORY_PATH) -> None:
        """Инициализирует объект для работы с файлом истории.

        :param path: Путь к <HISTORY_NAME>.xlsx.
        """
        self.path = path

    # ----------------------------- чтение -----------------------------
    def load_rows(self) -> List[HistoryRow]:
        """Читает <HISTORY_NAME>.xlsx и возвращает список записей истории.

        Поведение:
            - при отсутствии файла истории возвращает пустой список;
            - читает строки до первой «целиком пустой» (по Xlsx.is_blank_row);
            - пропускает строки без пути или с некорректной датой, пишет
              предупреждения в лог;
            - гарантирует, что каждая запись содержит нормализованный путь
              и дату (до минут).

        :return: Список HistoryRow в порядке следования строк.
        """
        # 1. Проверяем наличие файла истории.
        if not self.path.exists():
            log.info(
                "Файл истории не найден, история считается пустой: %s",
                self.path,
            )
            return []

        # 2. Открываем книгу в режиме только чтения.
        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        try:
            # 3. Проверяем наличие нужного листа.
            if SHEET_HISTORY not in wb.sheetnames:
                log.warning(
                    "Лист истории %s не найден в файле %s — история пустая",
                    SHEET_HISTORY,
                    self.path,
                )
                return []

            # Получаем объект листа из книги
            ws = wb[SHEET_HISTORY]
            rows: List[HistoryRow] = []

            # 4. Обходим строки с 2-й (шапка — 1-я) до первой полностью пустой.
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=2, values_only=True),
                start=2,
            ):
                if Xlsx.is_blank_row(row):
                    break

                # 4.1. Путь к модели — обязательный столбец A.
                path_str = Xlsx.cell(row, HISTORY_COL_RVT_PATH)
                if not path_str:
                    log.warning(
                        "Лист %s: строка %d пропущена — пустой путь к модели",
                        SHEET_HISTORY,
                        row_idx,
                    )
                    continue

                # 4.2. Дата в B: может быть datetime, строкой или
                #      Excel-сериалом.
                dt_val = None
                raw_dt = None
                if row and len(row) > HISTORY_COL_DATETIME:
                    raw_dt = row[HISTORY_COL_DATETIME]
                    dt_val = Xlsx.parse_datetime(raw_dt)

                if not dt_val:
                    # Строки с пустой/битой датой пропускаем.
                    log.warning(
                        "Лист %s: строка %d пропущена — некорректная дата: %r",
                        SHEET_HISTORY,
                        row_idx,
                        raw_dt,
                    )
                    continue

                # Нормализуем дату «до минут», чтобы история всегда
                # соответствовала контракту (как и mtime-функции).
                dt_val = dt_val.replace(second=0, microsecond=0)

                rows.append((path_str, dt_val))

            # 5. Логируем краткую сводку и возвращаем результат.
            log.info(
                "Загружено записей истории: %d (файл: %s)",
                len(rows),
                self.path.name,
            )

            return rows
        finally:
            wb.close()

    # ----------------------------- запись -----------------------------
    def save_rows(self, rows: Iterable[HistoryRow]) -> None:
        """Перезаписывает файл истории с полным пересозданием таблицы.

        :param rows: Итерация строк (path, datetime) для записи.
        """
        rows_list = list(rows)

        # открываем или создаем книгу
        wb = self._open_or_create_workbook()
        try:
            ws = self._get_or_create_sheet(wb)

            # Шапка и базовое оформление
            self._ensure_header(ws)
            self._format_sheet(ws)

            # Очищаем старые строки данных (если они есть).
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            # Записываем новые строки, получаем последнюю занятую строку.
            last_row = self._write_rows(ws, rows_list)
            end_row = max(last_row, 2)
            ref = f"A1:B{end_row}"

            # Пересоздаём Excel-таблицу и автофильтр.
            self._recreate_table(ws, ref)

            wb.save(self.path)
        finally:
            wb.close()

    # ----------------------- служебные методы -----------------------
    def _open_or_create_workbook(self) -> Workbook:
        """Открывает существующий файл или создаёт новый.

        Гарантирует существование родительской директории.

        :return: Открытая книга Excel.
        """
        ensure_dir(self.path.parent)

        if self.path.exists():
            try:
                return openpyxl.load_workbook(self.path)
            except Exception:
                log.warning(
                    "Не удалось открыть файл истории %s для перезаписи, "
                    "будет создан новый файл",
                    self.path,
                )

        # Новый файл: создаём книгу и переименовываем активный лист.
        wb = Workbook()
        cast(Worksheet, wb.active).title = SHEET_HISTORY

        return wb

    @staticmethod
    def _get_or_create_sheet(wb: Workbook) -> Worksheet:
        """Возвращает лист SHEET_HISTORY или создаёт новый с таким именем.

        :param wb: Открытая книга Excel.
        :return: Лист с именем SHEET_HISTORY.
        """
        if SHEET_HISTORY in wb.sheetnames:
            return cast(Worksheet, wb[SHEET_HISTORY])
        return wb.create_sheet(SHEET_HISTORY)

    # -------------------------- оформление листа --------------------------
    @staticmethod
    def _ensure_header(ws: Worksheet) -> None:
        """Создаёт шапку таблицы (ячейки A1:B1).

        Заголовки берутся из config.excel:
            - HISTORY_HDR_COL1 — путь к модели;
            - HISTORY_HDR_COL2 — дата модификации RVT на момент
              последней выгрузки.

        :param ws: Лист Excel, в который записывается шапка таблицы.
        """
        header_font = Font(
            color=HEADER_FONT_COLOR,
            bold=True,
            size=HEADER_FONT_SIZE,
        )
        headers = [HISTORY_HDR_COL1, HISTORY_HDR_COL2]

        # Формируем шапку по списку заголовков
        for i, text in enumerate(headers, start=1):
            cell = ws.cell(1, i, text)
            cell.alignment = Alignment(horizontal=ALIGN_HEADER)
            cell.font = header_font

    @staticmethod
    def _format_sheet(ws: Worksheet) -> None:
        """Настраивает ширину столбцов.

        :param ws: Лист Excel, для которого задаются ширины столбцов.
        """
        ws.column_dimensions["A"].width = COL_WIDTH_PATH
        ws.column_dimensions["B"].width = COL_WIDTH_DATE

    # ----------------------------- запись строк -----------------------------
    @staticmethod
    def _write_rows(ws: Worksheet, rows: Iterable[HistoryRow]) -> int:
        """Записывает строки и возвращает индекс последней заполненной строки.

        Если записей нет, создаёт скелет второй строки (A2:B2), чтобы
        таблица имела корректный диапазон даже при пустой истории.

        :param ws: Лист Excel, в который выполняется запись.
        :param rows: Итерация строк истории для записи.
        :return: Индекс последней заполненной строки.
        """
        row_idx = 2

        for path_str, dt in rows:
            HistoryXlsxIO._write_row(ws, row_idx, path_str, dt)
            row_idx += 1

        # Если не было ни одной строки — создаём пустую строку A2:B2.
        if row_idx == 2:
            HistoryXlsxIO._write_row(ws, 2, "", None)
            return 2

        return row_idx - 1

    @staticmethod
    def _write_row(
        ws: Worksheet,
        row_idx: int,
        path_str: Optional[str],
        dt: Optional[datetime],
    ) -> None:
        """Создаёт и оформляет одну строку истории.

        :param ws: Лист Excel, где создаются ячейки.
        :param row_idx: Индекс строки (2, 3, ...).
        :param path_str: Путь к модели или пустая строка для «скелета».
        :param dt: Дата модификации RVT или None для «скелета».
        """
        # Столбец A — путь к файлу.
        cell_path = ws.cell(
            row=row_idx,
            column=HISTORY_COL_RVT_PATH + 1,
            value=path_str,
        )
        cell_path.alignment = Alignment(horizontal=ALIGN_CELL)

        # Столбец B — дата выгрузки.
        cell_dt = ws.cell(
            row=row_idx,
            column=HISTORY_COL_DATETIME + 1,
            value=dt,
        )
        cell_dt.number_format = FORMAT_DATETIME_EXCEL
        cell_dt.alignment = Alignment(horizontal=ALIGN_CELL)

    # ------------------------- таблица и автофильтр -------------------------

    @staticmethod
    def _recreate_table(ws: Worksheet, ref: str) -> None:
        """Создаёт таблицу HISTORY_TBL_NAME с корректным ref и автофильтром.

        Перед созданием:
            - очищает ws.tables (dict) и ws._tables (set/list), чтобы
              избавиться от старых определений таблиц (защита от битого XML).

        :param ws: Лист, в котором создаётся таблица.
        :param ref: Диапазон (A1:B<N>) для таблицы и автофильтра.
        """
        # В разных версиях openpyxl таблицы могут храниться в .tables (dict)
        tables_obj = getattr(ws, "tables", None)
        if isinstance(tables_obj, dict):
            tables_obj.clear()

        # ...и/или во внутреннем поле _tables (set/list/tuple)
        tables_set = getattr(ws, "_tables", None)
        if tables_set:
            for tbl in list(tables_set):
                if hasattr(tables_set, "discard"):
                    tables_set.discard(tbl)
                elif hasattr(tables_set, "remove"):
                    try:
                        tables_set.remove(tbl)
                    except ValueError:
                        # Таблица могла уже исчезнуть — это не критично.
                        pass

        # Создаём новую таблицу поверх актуального диапазона ref.
        tbl = Table(displayName=HISTORY_TBL_NAME, ref=ref)
        tbl.tableColumns = [
            TableColumn(id=1, name=HISTORY_HDR_COL1),
            TableColumn(id=2, name=HISTORY_HDR_COL2),
        ]
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)

        # Настройка автофильтра. Сломаться не критично, поэтому try/except.
        try:
            if getattr(tbl, "autoFilter", None) is None:
                tbl.autoFilter = AutoFilter(ref)
            else:
                tbl.autoFilter.ref = ref  # type: ignore[attr-defined]
        except Exception as exc:
            log.debug(
                "Не удалось настроить автофильтр для таблицы history (%s): %s",
                ref,
                exc,
            )
