# -*- coding: utf-8 -*-
"""Загрузка <MANAGE_NAME>.xlsx → список RevitModel и ignore-set.

Назначение:
    - Прочитать Excel-файл <MANAGE_NAME>.xlsx и превратить его содержимое в
      структуру данных для дальнейшего экспорта:
        * список моделей (RevitModel),
        * множество путей-исключений (ignore),
        * список предупреждений по mtime.


Контракты:
    - Используются листы SHEET_PATH и SHEET_IGNORE (имена заданы в config).
    - Строки листов читаются последовательно до первой полностью пустой.
    - Столбцы интерпретируются согласно индексам в config.excel.
    - JSON-конфигурация маппинга и файл сопоставления категорий
      (family-mapping txt) обязаны существовать на диске:
        * при отсутствии любого из них выбрасывается ValueError.
    - Папки для выгрузки IFC создаются при необходимости (ensure_dir).
    - Время модификации RVT нормализуется до минут (file_mtime_minute).

Особенности:
    - Глобальный флаг FLAG_UNMAPPED управляет секцией выгрузки без маппинга
      (поля out_nomap_dir и nomap_json могут быть None).
    - Модуль ничего не пишет в файлы сам по себе — он только формирует
      в памяти структуру данных для последующих шагов.
"""
import logging
import openpyxl
from pathlib import Path
from dataclasses import dataclass
from typing import Set, List, Optional, Iterable

from config import (
    SHEET_PATH,
    LOGGER_NAME,
    MANAGE_NAME,
    MANAGE_PATH,
    SHEET_IGNORE,
    FLAG_UNMAPPED,
    DIR_EXPORT_CONFIG,
    DIR_MAPPING_COMMON,
    DIR_MAPPING_LAYERS,
    JSON_CONFIG_FILENAME,
)
from config.excel import (
    MANAGE_COL_RVT_DIR,
    MANAGE_COL_OUT_MAP,
    MANAGE_COL_MAP_DIR,
    MANAGE_COL_OUT_NOMAP,
    MANAGE_COL_FAMILY_MAP,
    MANAGE_COL_NOMAP_NAME,
    MANAGE_IGNORE_COL_PATH,
)

from core.models import RevitModel

from utils.xlsx_helpers import Xlsx
from utils.files import ensure_ext, is_pure_rvt
from utils.fs import ensure_dir, file_mtime_minute

# Модульный логгер: наследует настройки от "export_ifc"
log = logging.getLogger(f"{LOGGER_NAME}.manage")


class ManageDataLoader:
    """Загружает <MANAGE_NAME>.xlsx и формирует данные для экспорта.

    Состояние/результат:
        - models       — список RevitModel (по всем найденным .rvt);
        - ignore       — множество путей-исключений (строки, как в Excel);
        - models_mtime — сообщения о недоступном/пропущенном mtime.

    Правила:
        - Данные для моделей берутся с листа SHEET_PATH, игнор —
          с SHEET_IGNORE.
        - Чтение выполняется построчно до первой пустой строки.
        - Время модификации RVT нормализуется до минут.
    """

    def __init__(self, manage_path: Path = MANAGE_PATH) -> None:
        """Инициализирует загрузчик и сразу читает <MANAGE_NAME>.xlsx.

        :param manage_path: Путь к <MANAGE_NAME>.xlsx (по умолчанию —
                            из config).
        """
        self.manage_path = manage_path

        # Список моделей для выгрузки
        self.models: List[RevitModel] = []
        # Множество путей-исключений (строки, как в листе SHEET_IGNORE)
        self.ignore: Set[str] = set()
        # Сообщения о недоступном/пропущенном mtime
        self.models_mtime: List[str] = []

        # Проверка на наличие файла <MANAGE_NAME>.xlsx — без него
        # идти дальше нет смысла.
        if not self.manage_path.exists():
            raise FileNotFoundError(
                f"Не найден файл Excel с данными о моделях: {self.manage_path}"
            )

        log.info(
            "Чтение конфигурации из %s.xlsx: %s",
            MANAGE_NAME,
            self.manage_path
        )
        self._load()

    # ----------------------------- чтение -----------------------------
    def _load(self) -> None:
        """Загружает модели и игнор-лист из <MANAGE_NAME>.xlsx.

        Итого: заполняет self.models, self.ignore, self.models_mtime.
        """
        wb = openpyxl.load_workbook(
            self.manage_path,
            read_only=True,
            data_only=True,
        )
        try:
            self._read_path_sheet(wb)
            self._read_ignore(wb)
        finally:
            # Важно явно закрывать workbook, особенно в read_only-режиме.
            wb.close()

        log.info(
            "Загрузка %s.xlsx завершена: моделей=%d, "
            "ignore=%d, mtime_issues=%d",
            MANAGE_NAME,
            len(self.models),
            len(self.ignore),
            len(self.models_mtime),
        )

    # --------------------------- внутренний ---------------------------
    def _read_path_sheet(self, wb: openpyxl.Workbook) -> None:
        """Разбирает лист SHEET_PATH и собирает список моделей.

        :param wb: Открытая книга openpyxl.
        """
        if SHEET_PATH not in wb.sheetnames:
            # Без листа с путями сформировать модели невозможно —
            # просто выходим.
            log.warning(
                "Лист с путями (SHEET_PATH=%s) не найден в %s.xlsx",
                SHEET_PATH,
                MANAGE_NAME,
            )
            return

        # Получаем объект листа из книги
        ws = wb[SHEET_PATH]

        # Множество уже встреченных конфигов строк, чтобы не плодить
        # дубли задач
        seen_cfg: set[_RowCfg] = set()

        # Идём построчно, начиная со 2-й строки (1-я — шапка).
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            # Первая полностью пустая строка — сигнал остановки:
            # дальше данных по договорённости быть не должно.
            if Xlsx.is_blank_row(row):
                break

            # Парсим строку в нормализованный конфиг _RowCfg.
            cfg = self._parse_row(row)
            # Если обязательных данных нет — строка пропускается.
            if not cfg:
                log.warning(
                    "Лист %s: строка %d пропущена из-за неполных или "
                    "некорректных данных",
                    SHEET_PATH,
                    row_idx,
                )
                continue

            # Защита от дублирующих строк <MANAGE_NAME>.xlsx:
            # строки с одинаковым _RowCfg считаются дубликатами и пропускаются.
            if cfg in seen_cfg:
                log.warning(
                    "Лист %s: строка %d пропущена из-за "
                    "дублирования конфигурации",
                    SHEET_PATH,
                    row_idx,
                )
                continue
            # Сохраняем конфиг как уже обработанный
            seen_cfg.add(cfg)

            # Гарантируем существование выходных папок перед сбором задач.
            self._prepare_output_dirs(cfg)

            # Перебираем только «чистые» .rvt (без временных/копий).
            for rvt in self._iter_rvt_files(cfg.rvt_dir):
                # Нормализуем mtime до минут.
                mtime = file_mtime_minute(rvt)
                if not mtime:
                    # Если не удалось прочитать время модификации —
                    # сохраняем сообщение и пропускаем модель.
                    self.models_mtime.append(
                        f"{rvt} — не удалось определить время модификации"
                    )
                    continue

                # Собираем RevitModel — это ключевая структура,
                # которая дальше пойдёт в Orchestrator/History/IFCChecker.
                self.models.append(
                    RevitModel(
                        rvt_path=rvt,
                        last_modified=mtime,
                        output_dir_mapping=cfg.out_map_dir,
                        output_dir_nomap=cfg.out_nomap_dir,
                        mapping_json=cfg.mapping_json,
                        nomap_json=cfg.nomap_json,
                        family_mapping_file=cfg.family_mapping_file,
                    )
                )

    def _parse_row(self, row: tuple[object, ...]) -> Optional["_RowCfg"]:
        """Преобразует строку Excel в нормализованный конфиг.

        Назначение:
            - извлечь из строки пути и имена файлов;
            - проверить наличие обязательных полей;
            - подготовить нормализованный конфиг для дальнейшей обработки.

        :param row: Кортеж значений одной строки (values_only=True).
        :return: Экземпляр _RowCfg или None, если данные неполные
                 (строка тихо игнорируется).
        """

        def opt_path(idx: int) -> Optional[Path]:
            """Возвращает абсолютный Path из ячейки, если в ней
            непустое значение.

            :param idx: Индекс столбца в строке Excel.
            :return: Абсолютный Path, если в ячейке есть непустая
                     строка/значение и она интерпретируется как
                     абсолютный путь; иначе None.
            """
            # Xlsx.cell: безопасно берёт ячейку (обрезка/strip/None)
            s = Xlsx.cell(row, idx)
            if not s:
                return None

            p = Path(s)

            # Принимаем только абсолютные пути:
            # "C:\\...", "\\\\server\\share\\..." и т.п.
            if not p.is_absolute():
                return None

            return p

        def required_path(idx: int) -> Optional[Path]:
            """Возвращает путь к существующей директории из ячейки.

            :param idx: Индекс столбца в строке Excel.
            :return: Path, если в ячейке есть значение и соответствующая
                     директория существует; иначе None.
            """
            p = opt_path(idx)
            return p if p and p.exists() else None

        # ---- Обязательные поля ----
        # Папка-источник с .rvt (A).
        rvt_dir = required_path(MANAGE_COL_RVT_DIR)
        # Целевая папка для mapped-выгрузок (B).
        # Существование не проверяем — создадим при подготовке.
        out_map_dir = opt_path(MANAGE_COL_OUT_MAP)
        # Папка с JSON-конфигом маппинга (C) →
        # внутри ожидается JSON_CONFIG_FILENAME.
        map_dir = opt_path(MANAGE_COL_MAP_DIR)

        # Если какое-то из обязательных полей отсутствует —
        # строка игнорируется.
        if not (rvt_dir and out_map_dir and map_dir):
            # Обязательные пути не заданы/некорректны — вызывающий код
            # (_read_path_sheet) залогирует пропуск строки с указанием
            # листа и номера.
            return None

        mapping_json = map_dir / JSON_CONFIG_FILENAME
        # ---- FAIL-FAST: файл маппинга обязан существовать ----
        _ensure_exists(
            mapping_json, "файл JSON настроек выгрузки с маппингом"
        )

        # Имя txt-файла сопоставления категорий (D) →
        # лежит в DIR_MAPPING_LAYERS под DIR_EXPORT_CONFIG.
        fam = Xlsx.cell(row, MANAGE_COL_FAMILY_MAP)
        if not fam:
            # Без файла маппинга категорий строка не берется.
            return None

        fam = ensure_ext(fam, ".txt")
        family_mapping_file = Path(
            DIR_EXPORT_CONFIG) / DIR_MAPPING_LAYERS / fam
        # ---- FAIL-FAST: файл маппинга категорий должен существовать ----
        _ensure_exists(family_mapping_file, "файл маппинга Revit категорий")

        # ---- Необязательная секция выгрузки без маппинга ----
        # Управляется глобальным флагом FLAG_UNMAPPED.
        out_nomap_dir: Optional[Path] = None
        nomap_json: Optional[Path] = None

        if FLAG_UNMAPPED:
            # Папка назначения (E) + имя json-конфига (F) из DIR_MAPPING_COMMON
            out_nomap_dir = opt_path(MANAGE_COL_OUT_NOMAP)
            nomap_name = Xlsx.cell(row, MANAGE_COL_NOMAP_NAME)

            has_dir = out_nomap_dir is not None
            has_json_name = bool(nomap_name)

            # Неполная конфигурация nomap-сценария:
            # либо заполнена только папка, либо только имя JSON.
            # Такое состояние лучше считать ошибкой конфигурации,
            # чтобы не получать рассинхрон между IFCChecker и pyRevit.
            if has_dir != has_json_name:
                log.warning(
                    "Неполная конфигурация выгрузки без маппинга: "
                    "нужно заполнить и папку назначения, и имя JSON-файла "
                    "либо оставить обе ячейки пустыми."
                )
                return None

            if nomap_name:
                nomap_name = ensure_ext(nomap_name, ".json")
                nomap_json = (
                    Path(DIR_EXPORT_CONFIG) / DIR_MAPPING_COMMON / nomap_name
                )
                _ensure_exists(
                    nomap_json,
                    "файл JSON настроек выгрузки без маппинга",
                )

        return _RowCfg(
            rvt_dir=rvt_dir,
            out_map_dir=out_map_dir,
            out_nomap_dir=out_nomap_dir,
            mapping_json=mapping_json,
            family_mapping_file=family_mapping_file,
            nomap_json=nomap_json,
        )

    def _prepare_output_dirs(self, cfg: "_RowCfg") -> None:
        """Создаёт целевые директории выгрузки по конфигу строки.

        :param cfg: Нормализованный конфиг строки (_RowCfg).
        """
        # ensure_dir — идемпотентная операция: создаст папку при отсутствии,
        # при наличии — тихо ничего не сделает.
        ensure_dir(cfg.out_map_dir)
        if FLAG_UNMAPPED and cfg.out_nomap_dir:
            ensure_dir(cfg.out_nomap_dir)

    def _iter_rvt_files(self, rvt_dir: Path) -> Iterable[Path]:
        """Итерирует только корректные .rvt-файлы в заданной папке.

        :param rvt_dir: Папка с исходными Revit-моделями.
        :return: Итератор по допустимым .rvt-файлам (без временных
                 и «копий»).
        """
        # Сортировка для детерминированного порядка экспорта.
        for p in sorted(rvt_dir.glob("*.rvt")):
            # is_pure_rvt — отсекает временные/копии/ненужные варианты.
            if is_pure_rvt(p):
                yield p

    def _read_ignore(self, wb: openpyxl.Workbook) -> None:
        """Читает лист SHEET_IGNORE и формирует множество путей-исключений.

        :param wb: Открытая книга openpyxl.
        """
        if SHEET_IGNORE not in wb.sheetnames:
            # Отсутствие ignore-листа — допустимая ситуация: просто нечего
            # игнорировать.
            log.warning(
                "Лист ignore (SHEET_IGNORE=%s) не найден в %s.xlsx",
                SHEET_IGNORE,
                MANAGE_NAME,
            )
            return

        # Получаем объект листа из книги
        ws = wb[SHEET_IGNORE]

        # Идём построчно, начиная со 2-й строки (1-я — шапка).
        for row in ws.iter_rows(min_row=2, values_only=True):
            if Xlsx.is_blank_row(row):
                break

            path_str = Xlsx.cell(row, MANAGE_IGNORE_COL_PATH)
            if path_str:
                # Путь сохраняем в виде строки, без нормализации:
                # это позволит точно сопоставлять с тем, что указано в Excel.
                self.ignore.add(path_str)


# ------------------ Дополнительные классы/функции ------------------
@dataclass(frozen=True)
class _RowCfg:
    """Нормализованные данные одной строки Excel листа SHEET_PATH.

    Назначение:
        - Представляет собой результат парсинга одной строки таблицы
          <MANAGE_NAME>.xlsx (лист SHEET_PATH) и служит промежуточной моделью
          между Excel-данными и RevitModel.
    """
    rvt_dir: Path
    out_map_dir: Path
    out_nomap_dir: Optional[Path]
    mapping_json: Path
    family_mapping_file: Path
    nomap_json: Optional[Path]


def _ensure_exists(path: Path, what: str) -> None:
    """Проверяет, что путь существует на диске, иначе ValueError.

    :param path: Путь к файлу или директории.
    :param what: Описание ресурса для сообщения об ошибке.
    """
    if not path.exists():
        raise FileNotFoundError(f"не найден {what}: {path}")
