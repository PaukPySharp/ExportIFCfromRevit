# -*- coding: utf-8 -*-
import os
import csv
import openpyxl

from const.constants import NAME_FOLDER_NOT_MAPPING


def create_folder_not_mapping(data_about_rvt_files: list[dict]) -> None:
    # получаем список папок главной директории
    paths_to_main_folder = set(i['Путь_до_глав_папки']
                               for i in data_about_rvt_files)
    for path in paths_to_main_folder:
        path = fr'{path}\{NAME_FOLDER_NOT_MAPPING}'
        os.makedirs(path, exist_ok=True)
    return None


def create_task_files(dict_path_for_tasks: dict,
                      models_dict_path: dict) -> None:
    # запускаем цикл по словарю с путями до заданий
    for year in dict_path_for_tasks:
        path_task_file = models_dict_path[year]
        with open(path_task_file, 'w', encoding='utf-8') as task:
            for data, _ in dict_path_for_tasks[year]:
                task.write(f'{data}\n')


def create_file_with_path_and_mapping(data_from_tmp_file: list,
                                      path_tmp_file: str) -> None:
    # создаем tmp файл
    with open(path_tmp_file, 'w', encoding='utf-8', newline='') as tmp:
        writer = csv.writer(tmp, delimiter=';', quoting=csv.QUOTE_NONNUMERIC)
        # записываем данные с путями и папки маппинга
        writer.writerows(data_from_tmp_file)


def record_history_rvt_files(workbook: openpyxl.Workbook,
                             union_new_with_history_data: list) -> None:
    sheet_with_history = workbook['History']
    # формирует список кортежей формата
    # (путь до файла, дата и время его изменения)
    new_history_data = [(data_file['Путь_до_файла'],
                         data_file['ДатаВремя'])
                        for data_file in union_new_with_history_data]
    # удаление старых данных из таблицы со второй строки по двум столбцам
    for num_row in range(2, sheet_with_history.max_row + 1):
        sheet_with_history.cell(
            row=num_row, column=1).value = None
        sheet_with_history.cell(
            row=num_row, column=2).value = None
    # запись новых данных в таблицу со второй строки по двум столбцам
    for num_row, new_data in enumerate(new_history_data, 2):
        sheet_with_history.cell(
            row=num_row, column=1).value = new_data[0]
        sheet_with_history.cell(
            row=num_row, column=2).value = new_data[1]
