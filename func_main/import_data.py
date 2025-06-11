# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path
from datetime import datetime

from const.constants import FORMAT_DATETIME

from func_main.conversion import datetime_file_modification


def search_rvt_files(workbook: openpyxl.Workbook) -> list[dict]:
    sheet_with_folders = workbook['Path']  # получаем лист Path
    # создаем список для добавления в него путей до папок с rvt и шаблонами маппирования
    path_project_mapping = []
    # запускаем цикл со второй строки таблицы, чтобы пропустить шапку
    for num_row in range(2, sheet_with_folders.max_row + 1):
        # ячейка с путем до папки с rvt
        path = sheet_with_folders.cell(
            row=num_row, column=1).value
        # ячейка с названием проекта
        project = sheet_with_folders.cell(
            row=num_row, column=2).value
        # ячейка с названием папки с шаблоном маппирования
        mapping = sheet_with_folders.cell(
            row=num_row, column=3).value
        # объединяем данные
        tuple_path_and_mapping = (path, project, mapping)
        # проверка на пустые строки, если они начались прерываем цикл
        if any(map(
                lambda x: x.isspace()
                if isinstance(x, str)
                else x is None,
                tuple_path_and_mapping)):
            break
        else:
            # сцепляем полный путь до папки маппирования
            tuple_path_and_mapping = (path, fr'{project}\{mapping}')
            path_project_mapping.append(tuple_path_and_mapping)

    # создаем список для добавления в него данных об rvt файлах
    data_about_rvt_files = []
    for path_to_rvt, folder_mapping in path_project_mapping:
        # получаем все пути до файлов rvt из корня папки
        all_files = [i for i in Path(path_to_rvt).glob('*.rvt')
                     if list(map(str.lower, i.suffixes)) == ['.rvt']]
        # если файлы есть в папке, то добавляем данные о них в список
        if all_files:
            for path_to_file in all_files:
                data_about_rvt_files.append(
                    {
                        'Путь_до_файла': path_to_file,
                        'Путь_до_глав_папки': path_to_file.parent.parent,
                        'ИмяФайла': path_to_file.stem,
                        'ШаблонМаппирования': folder_mapping,
                        'ДатаВремя': datetime_file_modification(path_to_file)
                    }
                )

    return data_about_rvt_files


def get_ignore_files(workbook: openpyxl.Workbook) -> list[str]:
    sheet_ignore = workbook['IgnoreList']  # получаем лист IgnoreList

    file_ignore_data = []
    for num_row in range(2, sheet_ignore.max_row + 1):
        # строка с путем до файла
        path_to_file = sheet_ignore.cell(
            row=num_row, column=1).value
        # проверка на пустые строки, если они начались прерываем цикл
        if (path_to_file.isspace()
            if isinstance(path_to_file, str)
                else path_to_file is None):
            break
        else:
            # запись данных в список для игнора
            file_ignore_data.append(path_to_file)

    return file_ignore_data


def get_history_rvt_files(workbook: openpyxl.Workbook) -> list[dict]:
    sheet_with_history = workbook['History']  # получаем лист History

    file_history_data = []
    for num_row in range(2, sheet_with_history.max_row + 1):
        # строка с путем до файла
        path_to_file = sheet_with_history.cell(
            row=num_row, column=1).value
        # строка с датой модификации файла
        file_datetime = sheet_with_history.cell(
            row=num_row, column=2).value
        # защита от не правильного чтения типа данных в ячейке
        file_datetime = (datetime.strptime(file_datetime, FORMAT_DATETIME)
                         if isinstance(file_datetime, str)
                         else file_datetime)
        # проверка на пустые строки, если они начались прерываем цикл
        if any(i.isspace()
               if isinstance(i, str)
               else i is None
               for i in (path_to_file, file_datetime)):
            break
        else:
            # запись данных о прошлом состоянии файла в список
            file_history_data.append(
                {
                    'Путь_до_файла': path_to_file,
                    'ДатаВремя': file_datetime
                }
            )

    return file_history_data


def get_ifc_from_main_folder(data_about_rvt_files: list[dict]) -> dict:
    # получаем список папок где могут уже находится ifc
    paths_to_main_folder = set(i['Путь_до_глав_папки']
                               for i in data_about_rvt_files)
    # словарь вида {путь до папки: список путей до ifc}
    all_files_ifc = {}
    for path in paths_to_main_folder:
        all_files_ifc[path] = [i for i in Path(path).glob('*.ifc')]

    return all_files_ifc
